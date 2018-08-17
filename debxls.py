# 2018-08-12;kv;changed import because we now process xlsx and xls files
#from openpyxl import load_workbook
import openpyxl

# 2018-08-12;kv;added this routine to process xls files
import xlrd

import argparse
import sys
import datetime
import re
import glob

# execution example:
# python debxls.py FPELL 01-01-2017 12-31-2017
#

# set the appversion
AppVersion   = '1.04'

# filter values (used if we were not using command line args) - no longer required
GrantType    = 'FPELL'
startdatestr = '01-01-2017'
enddatestr   = '12-31-2017'

# defined output filenames
csvout_dumpall = 'DumpAllData.csv'


# set up at the beginning - the columns that we are reading in
xls_columns = [
    'Notused',   # index 0 is not used - so we just fill this in
    'Date',
    'Owed',
    'Paid',
    'Message',
    'Balance',
    'PastDue',
]

# set up at the beginning - the columns that we are outputting
xls_columns_out = [
    'filename',
    'worksheet',
    'row',
    'Date',
    'Owed',
    'Paid',
    'Message',
    'Balance',
    'PastDue',
    'GrantType',
    'Warning'
]

# input date format that we are expecting
input_date_format = '%m-%d-%Y'
low_date_string   = '01-01-2000'

# capture the current date time
now = datetime.datetime.now()

# ----------------- functions to start --------------------------------------------------------------

# this routine dumps out all the records
def dumpAllRecords(csvout_dumpall, xls_columns_out, xlsdata):
    # create the output file
    csvout = open(csvout_dumpall, 'w')
    
    # create the headers
    #print('--------------------------------------------------------')
    csvout.write(','.join(xls_columns_out))
    csvout.write("\n")
    
    # now create an output that is comma delimited
    for rec in xlsdata:
        newrec = []
        for colname in xls_columns_out:
            newrec.append( rec[colname] )

        # and then join/output this record
        csvout.write(','.join(newrec))
        csvout.write("\n")
            
    # close the output file
    csvout.close()
            
    # display the message
    print('Output file created:', csvout_dumpall)
    
# this routine shows all the XLS files that don't have a matching XLSX file - warnings
def xlsNoxlsxWarning( xlsfilelist, xlsxfilelist):
    
    # check for unprocessed files that are xls not xlsx
    unprocessed = []
    for xlsfile in xlsfilelist:
        # add the x to the filename
        xlsxfile = xlsfile + 'x'
        # check to see if this file exists - not exist capture the information
        if xlsxfile not in xlsxfilelist:
            unprocessed.append(xlsfile)

    # check - and if the lenght is not zero
    if len(unprocessed) > 0:
        print('The following XLS files were not processed - we can only process XLSX files - please convert or remove:')
        for file in unprocessed:
            print('XLS_file:', file)


# -----------------------------------------------------------------------------------------------            


# routine used to check that an entered string meets a date/time format and is a valid date
def valid_date(s):
    try:
        if datetime.datetime.strptime(s, input_date_format):
            return s
    except ValueError:
        msg = "Not a valid date: '{0}'.".format(s)
        raise argparse.ArgumentTypeError(msg)


# -----------------------------------------------------------------------------------------------            

    
# command line processor
parser = argparse.ArgumentParser()
parser.add_argument("GrantType", help="GrantType the string you use here", choices=['REGFEE','KIT','FPELL','FDSL-U','FDSL-S','TITLE-IV'])
parser.add_argument("StartDate", help="StartDate in MM-DD-YYYY format", type=valid_date)
parser.add_argument("EndDate", help="EndDate in MM-DD-YYYY format", type=valid_date)
args = parser.parse_args()

# pull the command line options
GrantType    = args.GrantType
startdatestr = args.StartDate
enddatestr   = args.EndDate

#convert to date fields
startdate = datetime.datetime.strptime(startdatestr, input_date_format)
enddate   = datetime.datetime.strptime(enddatestr,   input_date_format)

# output filenames - calculated based on input values
csvout_filtered = 'Filter-' + GrantType + '-' + startdatestr + '-' + enddatestr + '.csv'


# read the file list using glob to load in xlsx and then xls files
xlsfilelist = glob.glob('./*.xlsx')
xlsfilelist.extend( glob.glob('./*.xls') )

# uncomment when debugging xls only files
# xlsfilelist = glob.glob('./*.xls')

# debugging
# file list
# xlsxfilelist = ['./Actual Dispersement record.xlsx']
# print( xlsxfilelist )
# sys.exit()

# 2018-08-12;kv;this code removed because we can now process xlsx and xls files
# display the warnings about xls with no xlsx
#xlsNoxlsxWarning( xlsfilelist, xlsxfilelist)

# create array that holds all the read in data
xlsdata = []

# create array that holds the exception data
xlsdataerror = []

#--------------------------------------------------------------------------------------------

# loop through the files of interest
for xlsfilename in xlsfilelist:

    # check the filetype to see if it starts with ~, if so skip filename
    if re.match('\.\\\\~', xlsfilename):
        print('Skipped filename...:', xlsfilename)
        continue

    
    # determine what filetype we have here
    xlsxfiletype = re.search('.xlsx$', xlsfilename)

    # debugging
    #print('xlsfilename:',xlsfilename)
    #print('xlsxfiletype:',xlsxfiletype)
    
    # Load in the workbook (set the data_only=True flag to get the value on the formula)
    if xlsxfiletype:
        # XLSX file
        wb = openpyxl.load_workbook(xlsfilename, data_only=True)
        sheetNames = wb.sheetnames
    else:
        # XLS file
        wb = xlrd.open_workbook(xlsfilename)
        sheetNames = wb.sheet_names()


    # get the list of sheets that need to process
    for sheetName in sheetNames:
        # create a workbook sheet object - using the name to get to the right sheet
        if xlsxfiletype:
            s = wb[sheetName]
        else:
            s = wb.sheet_by_name(sheetName)
        
        # grab the sheet title - not sure i need this - that is already in sheetName
        if xlsxfiletype:
            sheettitle = s.title
            sheetmaxrow = s.max_row
            sheetmaxcol = s.max_column
        else:
            sheettitle = s.name
            sheetmaxrow = s.nrows
            sheetmaxcol = s.ncols
        
        #### Find the header row - need ot define the column and the value
        
        # the column that has this value
        column = 1
        
        # find the row that has the headers
        for row_header in range(1,6):
            # get cell value
            if xlsxfiletype:
                cValue = s.cell(row=row_header, column=column).value
            else:
                cValue = s.cell(row_header, column-1).value

            # check to see if this the header row
            if cValue == xls_columns[column]:
                # this is the header row - validate a few more fields
                column = 5
                # get cell value again
                if xlsxfiletype:
                    cValue = s.cell(row=row_header, column=column).value
                else:
                    cValue = s.cell(row_header, column-1).value
                # check value again
                if cValue != xls_columns[column]:
                    # we have a problem - so display message and exit
                    print('Column[', column, '] should be (', xls_columns[column], ') but is:', cValue)
                    print('Workbook:', xlsfilename)
                    print('Sheetname:', sheetName)
                    print('Row:', row_header)
                    print('Exit and fix or remove XLSX')
                    sys.exit(1)
                # did not fail - so break out we have the header
                break
        
        # debugging - print out what we found
        #print ('found the matching column:', row_header, ':', column)

        # create starting comparison date
        lastDate = datetime.datetime.strptime(low_date_string, input_date_format)
        
        # pull in all the data from this sheet that we are interested in 
        for row in range(row_header+1, sheetmaxrow):
            # create a new record to hold this rows data
            rec = {}

            # fill in the major attributes
            rec['worksheet'] = sheetName
            rec['filename'] = xlsfilename
            rec['row'] = str(row)

            # go through the columns of this row
            for col in range(1,7):
                # now populate the record
                if xlsxfiletype:
                    rec[xls_columns[col]] = s.cell(row=row, column=col).value
                else:
                    rec[xls_columns[col]] = s.cell(row, col-1).value

                # DATE - special row processing logic
                if xls_columns[col] == 'Date':
                    # debugging
                    #print('type:rec[date]:', type(rec['Date']), '-value:', rec['Date'])

                    # convert to datetime if the value is float (usually used for xls files)
                    if isinstance(rec['Date'],float):
                        # use the routine to convert the field
                        rec['Date'] = xlrd.xldate.xldate_as_datetime(rec['Date'],0)
                        # debugging
                        #print('type:rec[date]2:', type(rec['Date']), '-value:', rec['Date'])
                    
                    # create a copy to use at other time
                    rec['DateDate'] = rec['Date']

                    # if the date field is populated
                    # test to see if the date is a string type
                    if isinstance(rec['Date'],str):
                        # date should not be a string field - warning message
                        if 'Warning' in rec.keys():
                            rec['Warning'] += ':' + 'Date-string'
                        else:
                            rec['Warning'] = 'Date-string'
                    elif isinstance(rec['Date'],int):
                        # date should not be a string field - warning message
                        if 'Warning' in rec.keys():
                            rec['Warning'] += ':' + 'Date-int'
                        else:
                            rec['Warning'] = 'Date-int'
                    elif rec['Date'] != None:
                        # if the current date field is less than the last value
                        if rec['Date'] < lastDate:
                            # add to the record a message
                            if 'Warning' in rec.keys():
                                rec['Warning'] += ':' + 'Date-earlier'
                            else:
                                rec['Warning'] = 'Date-earlier'
                        else:
                            # reset the lastDate
                            lastDate = rec['Date']
                    #else:
                        # debugging
                        # print('date is None')
                            
                # MESSAGE - special row processing logic
                if xls_columns[col] == 'Message':
                    # check for what grant type this should be
                    if rec['Message'] == None:
                        # message is blank
                        rec['GrantType'] = ''

                        # debugging
                        #print('message-blank-date-value:', rec['Date'])
                        
                        # now test if the date field is populated
                        if rec ['Date'] != 'None':
                            # debugging
                            # print('message-blank-date-not-none-date-value:', rec['Date'])
                            
                            # add to the record a message
                            if 'Warning' in rec.keys():
                                rec['Warning'] += ':' + 'Msg-blank'
                            else:
                                rec['Warning'] = 'Msg-blank'
                    elif not isinstance(rec['Message'],str):
                        # message is not string
                        rec['GrantType'] = ''

                        # message is not string - so we should not regex
                        if 'Warning' in rec.keys():
                            rec['Warning'] += ':' + 'Msg-NotString'
                        else:
                            rec['Warning'] = 'Msg-NotString'
                    # check for what grant type this should be
                    elif rec['Message'] == '':
                        # message is blank
                        rec['GrantType'] = ''

                        # debugging
                        #print('message-blank-date-value:', rec['Date'])
                        
                        # now test if the date field is populated
                        if rec ['Date'] != 'None':
                            # debugging
                            # print('message-blank-date-not-none-date-value:', rec['Date'])
                            
                            # add to the record a message
                            if 'Warning' in rec.keys():
                                rec['Warning'] += ':' + 'Msg-blank'
                            else:
                                rec['Warning'] = 'Msg-blank'
                    elif re.search('registration', rec['Message'], re.IGNORECASE):
                        # registration fee
                        rec['GrantType'] = 'REGFEE'
                    elif re.search('kit', rec['Message'], re.IGNORECASE):
                        # kit
                        rec['GrantType'] = 'KIT'
                    elif re.search('FPELL', rec['Message'], re.IGNORECASE):
                        # FPELL
                        rec['GrantType'] = 'FPELL'
                    elif re.search('FDSL-U', rec['Message'], re.IGNORECASE):
                        # FDSL-U
                        rec['GrantType'] = 'FDSL-U'
                    elif re.search('FDSL-S', rec['Message'], re.IGNORECASE):
                        # FDSL-S
                        rec['GrantType'] = 'FDSL-S'
                    elif re.search('title\s+iv', rec['Message'], re.IGNORECASE):
                        # Title IV
                        rec['GrantType'] = 'TITLE-IV'
                    else:
                        rec['GrantType'] = ''

                
                # PAID - special row processing logic
                if xls_columns[col] == 'Paid':
                    # check to see if this is None or Zero
                    if rec['Paid'] == 0 or rec['Paid'] == None:
                        # check to see if the Date field is type date and the date is earlier than today
                        if isinstance(rec['DateDate'], datetime.datetime):
                            # this is a date see if it is in the past
                            if rec['DateDate'] < now:
                                # add to the record a message
                                if 'Warning' in rec.keys():
                                    rec['Warning'] += ':' + 'PASTDUE'
                                else:
                                    rec['Warning'] = 'PASTDUE'
                                
                # convert values to string
                if isinstance(rec[xls_columns[col]], datetime.datetime):
                    #print(sheetName,':',row,':',col,':converted-field')
                    rec[xls_columns[col]] = rec[xls_columns[col]].strftime(input_date_format)
                else:
                    rec[xls_columns[col]] = str(rec[xls_columns[col]])

                # make sure there are no comma's in these strings
                rec[xls_columns[col]] = re.sub(',', ';', rec[xls_columns[col]])

            # check to see that the warning field is populated
            if 'Warning' not in rec.keys():
                rec['Warning'] = ''
                    
            # now show what we got
            #print('rec:',rec)
        
            # now add this record to the current array
            xlsdata.append(rec)

# we are done - print out all the records
#print('all-records:', xlsdata)

# now output all the datea
dumpAllRecords(csvout_dumpall, xls_columns_out, xlsdata)


# build the filter records
xlsfiltered = []
for rec in xlsdata:
    # check to see if this is the right granttype
    if rec['GrantType'] == GrantType:
        # check to see if this is a record of date type
        if isinstance(rec['DateDate'], datetime.datetime):
            # check to see if we are aligned to start date
            if rec['DateDate'] >= startdate:
                # check to see if we are aligned to end date
                if rec['DateDate'] <= enddate:
                    # append this record to the filtered array
                    xlsfiltered.append(rec)
        else:
            print('DateDate not date time but is:', type(rec['DateDate']))

# we have the filtered list dump it out
dumpAllRecords(csvout_filtered, xls_columns_out, xlsfiltered)

sys.exit()

# eof
